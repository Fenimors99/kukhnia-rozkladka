"""
rozkladka_core.py — shared parameterized logic for food schedule generation.
Cross-platform (macOS / Windows). No hardcoded paths, dates, or counts.
"""
import re
import os
from pathlib import Path
from datetime import datetime, timedelta
from itertools import groupby

import openpyxl
from num2words import num2words
from weasyprint import HTML as _WP_HTML


def _html_to_pdf(html: str, out_path: Path):
    _WP_HTML(string=html).write_pdf(str(out_path))


DAYS_ORDER = ['Понеділок', 'Вівторок', 'Середа', 'Четвер', "П'ятниця", 'Субота', 'Неділя']


# ── Formatting ────────────────────────────────────────────────────────────────

def fmt(v):
    if v is None or v == '':
        return ''
    if isinstance(v, float):
        if v == 0.0:
            return '0'
        s = f'{v:.3f}'.rstrip('0').rstrip('.')
        return s.replace('.', ',')
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


# ── Date utilities ────────────────────────────────────────────────────────────

def detect_dates_from_filename(path: str):
    """
    Try to extract week start date from xlsx filename.
    Example: '...06.04-12.04.2026...' → '06.04.2026'
    Returns 'DD.MM.YYYY' string or None.
    """
    name = Path(path).name
    m = re.search(r'(\d{2})\.(\d{2})[._-](\d{2})\.(\d{2})\.(\d{4})', name)
    if m:
        return f'{m.group(1)}.{m.group(2)}.{m.group(5)}'
    return None


def build_days_dates(start_date_str: str) -> dict:
    """Build {day_name: 'DD.MM.YYYY'} for 7 days from start_date_str 'DD.MM.YYYY'."""
    start = datetime.strptime(start_date_str, '%d.%m.%Y')
    return {DAYS_ORDER[i]: (start + timedelta(days=i)).strftime('%d.%m.%Y') for i in range(7)}


def date_range_label(start_date_str: str) -> str:
    """'06.04.2026' → '06.04–12.04.2026'"""
    start = datetime.strptime(start_date_str, '%d.%m.%Y')
    end = start + timedelta(days=6)
    return f"{start.strftime('%d.%m')}–{end.strftime('%d.%m.%Y')}"


# ── Column auto-detection ────────────────────────────────────────────────────

def detect_column_range(raw: list) -> tuple:
    """
    Auto-detect ingredient column range from raw xlsx data.
    Returns (ing_start, ing_end, total_col, meat_col).
    ing_start..ing_end is exclusive: range(ing_start, ing_end).
    """
    total_col = None
    meat_col = None

    # Scan header rows 0-13 for keyword hints
    for row_idx in range(min(14, len(raw))):
        for col_idx, cell in enumerate(raw[row_idx]):
            if not cell:
                continue
            s = str(cell).strip().lower()
            if total_col is None and 'загальна маса готової' in s:
                total_col = col_idx
            if meat_col is None and "м'ясних та рибних" in s:
                meat_col = col_idx
        if total_col is not None and meat_col is not None:
            break

    if total_col is not None and meat_col is not None:
        return 4, total_col, total_col, meat_col

    # Fallback: scan first "Усього за день" row — last 2 non-None cols are meat/total
    for row in raw[12:]:
        if len(row) < 2:
            continue
        col1 = str(row[1]).strip() if row[1] else ''
        if col1 == 'Усього за день':
            filled = [i for i in range(4, len(row)) if row[i] is not None]
            if len(filled) >= 2:
                return 4, filled[-2], filled[-2], filled[-1]

    return 4, 80, 80, 81  # hard fallback


# ── CSS ───────────────────────────────────────────────────────────────────────

_CSS_DAILY = """
@page { size: A3 landscape; margin: 4mm; }
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: Arial, sans-serif; font-size: 6pt; background: white; }
h2 { text-align: center; font-size: 8pt; font-weight: bold; margin-bottom: 2mm; }
table { border-collapse: collapse; width: 100%; table-layout: fixed; }
th, td { border: 0.4pt solid #000; padding: 1px 1px; text-align: center;
         vertical-align: middle; line-height: 1.2; font-size: 6pt; }
td.c-name { text-align: left; font-size: 8pt; padding-left: 3px; }
th.v-hdr { height: 42mm; padding: 0; overflow: hidden; }
th.ing-main-hdr { font-size: 6pt; font-weight: bold; padding: 2px; }
td.meal-cell { background: #e0e0e0; padding: 0; }
tr.total td { background: #e8e8e8; font-weight: bold; }
@media print { body { -webkit-print-color-adjust: exact; print-color-adjust: exact; } }
"""

_CSS_PERIOD = """
@page { size: A3 landscape; margin: 4mm; }
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: Arial, sans-serif; font-size: 6pt; background: white; }
h2 { text-align: center; font-size: 8pt; font-weight: bold; margin-bottom: 2mm; }
.doc-header {
    display: flex; justify-content: space-between;
    margin-bottom: 2mm; font-size: 6.5pt; line-height: 1.5;
}
.doc-header-left { flex: 1; }
.doc-header-right { text-align: right; min-width: 80mm; }
.doc-header-left .title { font-size: 9pt; font-weight: bold; }
.doc-header-right .zatv { font-size: 8pt; font-weight: bold; }
.doc-header .hint { font-size: 5.5pt; color: #555; }
.doc-footer { margin-top: 3mm; font-size: 6.5pt; line-height: 1.6; }
.doc-footer-row { display: flex; justify-content: space-between; margin-top: 1mm; }
.doc-footer-col { flex: 1; }
.doc-footer .hint { font-size: 5.5pt; color: #555; }
table { border-collapse: collapse; width: 100%; table-layout: fixed; }
th, td {
    border: 0.4pt solid #000; padding: 0px 2px; text-align: center;
    vertical-align: middle; line-height: 1.1; font-size: 5.5pt;
}
td.name, th.name { text-align: left; font-size: 6pt; padding-left: 3px; }
th.day-hdr { background: #d0d0d0; font-weight: bold; font-size: 7pt; }
th.period-hdr { background: #a0a0a0; font-weight: bold; font-size: 7pt; }
tr.param td { background: #f0f0f0; font-style: italic; }
tr.param td.name { font-style: italic; }
td.period-val { background: #e0e0e0; font-weight: bold; }
tr:nth-child(even) td:not(.period-val) { background: #fafafa; }
tr:nth-child(even) td.name { background: #fafafa; }
tr.param:nth-child(even) td { background: #f0f0f0; }
@media print { body { -webkit-print-color-adjust: exact; print-color-adjust: exact; } }
"""


# ── SVG vertical text helper ─────────────────────────────────────────────────

def _svg_vtext(text, h_mm=42, fontsize_pt=5, bold=True):
    """Return SVG element with text rotated bottom-to-top, fills parent width.
    Long text is word-wrapped into multiple lines spread across the cell width."""
    h = h_mm
    fs = round(fontsize_pt * 0.3528, 2)   # pt → mm (viewBox units)
    lh = round(fs * 1.35, 2)              # line height
    fw = 'bold' if bold else 'normal'

    # Estimate max chars per line based on available height
    max_chars = max(8, int(h_mm / max(0.1, fs * 0.55)))

    # Word-wrap
    words = text.split()
    lines, cur = [], ''
    for w in words:
        test = (cur + ' ' + w).strip()
        if cur and len(test) > max_chars:
            lines.append(cur)
            cur = w
        else:
            cur = test
    if cur:
        lines.append(cur)

    n = len(lines)
    parts = []
    for i, line in enumerate(lines):
        xi = round(5 + (i - (n - 1) / 2) * lh, 3)
        t = line.replace('&', '&amp;').replace('<', '&lt;')
        parts.append(
            f'<text x="{xi}" y="{h/2}" transform="rotate(-90 {xi} {h/2})" '
            f'text-anchor="middle" dominant-baseline="middle" '
            f'font-size="{fs}" font-family="Arial,sans-serif" font-weight="{fw}">'
            f'{t}</text>'
        )

    return (
        f'<svg xmlns="http://www.w3.org/2000/svg" width="100%" height="{h}mm" '
        f'viewBox="0 0 10 {h}" preserveAspectRatio="xMidYMid meet">'
        + ''.join(parts) +
        f'</svg>'
    )


# ── Daily PDF generation ──────────────────────────────────────────────────────

def generate_daily(xlsx_path: str, out_path: str, unit: str, start_date_str: str,
                   progress_cb=None):
    """
    Generate combined A3-landscape PDF (7 days, one page each) from xlsx.
    out_path: full path to output .pdf file.
    progress_cb: callable(str) for log output, or None.
    """
    def log(msg):
        if progress_cb:
            progress_cb(msg)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb['Аркуш1']
    raw = list(ws.iter_rows(values_only=True))

    ing_start, ing_end, total_col, meat_col = detect_column_range(raw)
    N_ING = ing_end - ing_start
    ING_NAMES = [
        str(raw[11][i]).strip() if i < len(raw[11]) and raw[11][i] else ''
        for i in range(ing_start, ing_end)
    ]

    DAYS_DATES = build_days_dates(start_date_str)
    range_label = date_range_label(start_date_str)

    days = {d: [] for d in DAYS_ORDER}
    cur_day = None
    cur_meal = None

    for row in raw[12:]:
        col0 = str(row[0]).strip() if row[0] else ''
        col1 = str(row[1]).strip() if row[1] else ''
        col2 = str(row[2]).strip() if row[2] else ''

        if col0:
            for d in DAYS_ORDER:
                if col0.startswith(d):
                    cur_day = d
                    break

        if col1 and col1 != 'Усього за день':
            cur_meal = col1

        if not cur_day:
            continue

        ings = [fmt(row[i]) if i < len(row) else '' for i in range(ing_start, ing_end)]
        tot  = fmt(row[total_col]) if total_col < len(row) else ''
        meat = fmt(row[meat_col])  if meat_col  < len(row) else ''

        if col1 == 'Усього за день':
            days[cur_day].append({
                'meal': '__total__', 'dish': 'Усього за день',
                'pct': '', 'total': tot, 'meat': meat, 'ings': ings,
            })
            continue

        if not col2:
            continue

        days[cur_day].append({
            'meal':  cur_meal or '',
            'dish':  col2,
            'pct':   str(row[3]).strip() if row[3] else '',
            'total': tot, 'meat': meat, 'ings': ings,
        })

    def _build_day(day_name, rows, inner_only=False):
        date = DAYS_DATES[day_name]
        total_row = next((r for r in rows if r['meal'] == '__total__'), None)
        data_rows = [r for r in rows if r['meal'] != '__total__']
        used_idx = [i for i in range(N_ING) if any(r['ings'][i] for r in data_rows)]

        name_w = 34
        avail = 408 - (8 + name_w + 6 + 11 + 11)
        ing_w = round(avail / len(used_idx), 1) if used_idx else 7
        title = f'РОЗКЛАДКА ПРОДУКТІВ — {day_name}, {date} — Підрозділ {unit}'

        if inner_only:
            L = [f'<h2>{title}</h2>', '<table>']
        else:
            L = [
                '<!DOCTYPE html><html lang="uk"><head>',
                '<meta charset="UTF-8">',
                f'<title>Розкладка — {day_name} {date}</title>',
                f'<style>{_CSS_DAILY}</style>',
                '</head><body>',
                f'<h2>{title}</h2>',
                '<table>',
            ]

        meal_groups = [(m, list(g)) for m, g in groupby(data_rows, key=lambda r: r['meal'])]

        L += ['<colgroup>',
              '<col style="width:8mm">',
              f'<col style="width:{name_w}mm">',
              '<col style="width:6mm">']
        for _ in used_idx:
            L.append(f'<col style="width:{ing_w}mm">')
        L += ['<col style="width:11mm">', '<col style="width:11mm">', '</colgroup>']

        L += ['<thead><tr>',
              f'<th class="v-hdr" rowspan="2">{_svg_vtext("Прийняття їжі")}</th>',
              f'<th class="v-hdr" rowspan="2">{_svg_vtext("Найменування страв")}</th>',
              f'<th class="v-hdr" rowspan="2">{_svg_vtext("% страви за типом")}</th>',
              f'<th class="ing-main-hdr" colspan="{len(used_idx)}">Найменування продуктів та маса їх в грамах на одну особу</th>',
              f'<th class="v-hdr" rowspan="2">{_svg_vtext("Загальна маса готової страви, г")}</th>',
              '<th class="v-hdr" rowspan="2">' + _svg_vtext("Маса м'ясних та рибних порцій, г") + '</th>',
              '</tr><tr>']
        for i in used_idx:
            L.append(f'<th class="v-hdr">{_svg_vtext(ING_NAMES[i])}</th>')
        L.append('</tr></thead><tbody>')

        for meal, dishes in meal_groups:
            for idx, r in enumerate(dishes):
                L.append('<tr>')
                if idx == 0:
                    meal_h = max(10, len(dishes) * 5)
                    L.append(f'<td class="meal-cell" rowspan="{len(dishes)}">{_svg_vtext(meal, h_mm=meal_h, fontsize_pt=7)}</td>')
                L.append(f'<td class="c-name">{r["dish"]}</td>')
                L.append(f'<td>{r["pct"]}</td>')
                for i in used_idx:
                    L.append(f'<td>{r["ings"][i]}</td>')
                L += [f'<td>{r["total"]}</td>', f'<td>{r["meat"]}</td>', '</tr>']

        if total_row:
            L += ['<tr class="total">',
                  '<td colspan="2" class="c-name">Усього за день</td>',
                  '<td></td>']
            for i in used_idx:
                L.append(f'<td>{total_row["ings"][i]}</td>')
            L += [f'<td>{total_row["total"]}</td>', f'<td>{total_row["meat"]}</td>', '</tr>']

        L.append('</tbody></table>')
        if not inner_only:
            L.append('</body></html>')
        return '\n'.join(L)

    for day in DAYS_ORDER:
        rows = days[day]
        if not rows:
            log(f'⚠️  {day}: немає рядків')
            continue
        used = len([i for i in range(N_ING) if any(r['ings'][i] for r in rows if r['meal'] != '__total__')])
        log(f'  {day} ({DAYS_DATES[day]}): {len(rows)-1} страв, {used} інгр.')

    css_combined = _CSS_DAILY + '\n.day-wrap{break-after:page;page-break-after:always}.day-wrap:last-child{break-after:avoid;page-break-after:avoid}'
    combined = [
        '<!DOCTYPE html><html lang="uk"><head>',
        '<meta charset="UTF-8">',
        f'<title>Розкладка продуктів — {range_label} — {unit}</title>',
        f'<style>{css_combined}</style>',
        '</head><body>',
    ]
    for day in DAYS_ORDER:
        if not days[day]:
            continue
        combined += ['<div class="day-wrap">', _build_day(day, days[day], inner_only=True), '</div>']
    combined.append('</body></html>')

    log('Конвертую в PDF…')
    _html_to_pdf('\n'.join(combined), out_path)
    log(f'✅  Збережено: {out_path.name}')


# ── Period summary HTML generation ───────────────────────────────────────────

def generate_period(xlsx_path: str, out_path: str, unit: str, start_date_str: str,
                    progress_cb=None):
    """
    Generate period-summary HTML (one table, all 7 days as columns).
    progress_cb: callable(str) or None.
    """
    def log(msg):
        if progress_cb:
            progress_cb(msg)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb['Аркуш1']
    raw = list(ws.iter_rows(values_only=True))

    ing_start, ing_end, total_col, meat_col = detect_column_range(raw)
    N_ING = ing_end - ing_start
    ING_NAMES = [
        str(raw[11][i]).strip() if i < len(raw[11]) and raw[11][i] else ''
        for i in range(ing_start, ing_end)
    ]

    start_dt = datetime.strptime(start_date_str, '%d.%m.%Y')
    range_label = date_range_label(start_date_str)
    DAYS = [(d, (start_dt + timedelta(days=i)).strftime('%d.%m')) for i, d in enumerate(DAYS_ORDER)]

    day_totals = []
    for row in raw:
        col1 = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        if col1 == 'Усього за день':
            day_totals.append({
                'ings':  [fmt(row[i]) if i < len(row) else '' for i in range(ing_start, ing_end)],
                'total': fmt(row[total_col]) if total_col < len(row) else '',
                'meat':  fmt(row[meat_col])  if meat_col  < len(row) else '',
            })

    period_total = None
    for row in raw:
        col0 = str(row[0]).strip() if row[0] else ''
        if col0.startswith('Усього за період'):
            period_total = {
                'ings':  [fmt(row[i]) if i < len(row) else '' for i in range(ing_start, ing_end)],
                'total': fmt(row[total_col]) if total_col < len(row) else '',
                'meat':  fmt(row[meat_col])  if meat_col  < len(row) else '',
            }
            break

    log(f"Знайдено 'Усього за день': {len(day_totals)} (очікується 7)")
    log(f"'Усього за період': {'знайдено' if period_total else 'НЕ ЗНАЙДЕНО'}")
    if len(day_totals) != 7:
        log('⚠️  Кількість днів не відповідає 7, перевір xlsx')

    all_cols = [d['ings'] for d in day_totals]
    if period_total:
        all_cols.append(period_total['ings'])
    used_idx = [i for i in range(N_ING) if any(col[i] for col in all_cols)]
    log(f'Інгредієнтів з даними: {len(used_idx)}')

    start_label = start_dt.strftime('%d.%m')
    end_label = (start_dt + timedelta(days=6)).strftime('%d.%m.%Y')

    L = [
        '<!DOCTYPE html><html lang="uk"><head>',
        '<meta charset="UTF-8">',
        f'<title>Розкладка — Усього за період {range_label}</title>',
        f'<style>{_CSS_PERIOD}</style>',
        '</head><body>',
        '<div class="doc-header">',
        '  <div class="doc-header-left">',
        '    <div class="title">РОЗКЛАДКА ПРОДУКТІВ</div>',
        '    <div>за нормами пайка</div>',
        '    <div>(на одну особу на добу)</div>',
        f'    <div><b>{unit}</b></div>',
        '    <div class="hint">(військова частина, підрозділ)</div>',
        f'    <div>на час з {start_label} по {end_label}</div>',
        '  </div>',
        '  <div class="doc-header-right">',
        '    <div class="zatv">ЗАТВЕРДЖУЮ</div>',
        '    <div>_______________________________</div>',
        '    <div class="hint">(посада)</div>',
        '    <div>_______________________________</div>',
        '    <div class="hint">(військове звання, підпис, прізвище)</div>',
        '    <div>"___" _______________ 20__ р.</div>',
        '  </div>',
        '</div>',
        f'<h2>Усього за період — {range_label} — Підрозділ {unit}</h2>',
        '<table>',
        '<colgroup>',
        '<col style="width:70mm">',
    ]
    for _ in DAYS:
        L.append('<col style="width:30mm">')
    L += ['<col style="width:35mm">', '</colgroup>', '<thead><tr>',
          '<th class="name">Найменування продуктів та маса їх в грамах на одну особу</th>']
    for name, date in DAYS:
        L.append(f'<th class="day-hdr">{name}<br>{date}.{start_dt.year}</th>')
    L += ['<th class="period-hdr">Усього за період</th>', '</tr></thead>', '<tbody>']

    for label, key in [('Загальна маса готової страви, г', 'total'),
                       ("Маса м'ясних та рибних порцій, г", 'meat')]:
        L.append(f'<tr class="param"><td class="name">{label}</td>')
        for d in range(len(day_totals)):
            L.append(f'<td>{day_totals[d][key]}</td>')
        L += [f'<td class="period-val">{period_total[key] if period_total else ""}</td>', '</tr>']

    for i in used_idx:
        L.append(f'<tr><td class="name">{ING_NAMES[i]}</td>')
        for d in range(len(day_totals)):
            L.append(f'<td>{day_totals[d]["ings"][i]}</td>')
        L += [f'<td class="period-val">{period_total["ings"][i] if period_total else ""}</td>', '</tr>']

    L += [
        '</tbody></table>',
        '<div class="doc-footer">',
        '  <div>Заступник командира військової частини з тилу (матеріально-технічного забезпечення)</div>',
        '  <div class="hint">(військове звання, підпис, прізвище)</div>',
        '  <div class="doc-footer-row" style="margin-top:2mm">',
        '    <div class="doc-footer-col">',
        '      <div>Начальник продовольчої служби</div>',
        '      <div class="hint">(військове звання, підпис, прізвище)</div>',
        '    </div>',
        '    <div class="doc-footer-col" style="text-align:center">',
        '      <div>Начальник медичної служби</div>',
        '      <div class="hint">(старший лікар)</div>',
        '    </div>',
        '    <div class="doc-footer-col" style="text-align:right">',
        '      <div>&nbsp;</div>',
        '      <div class="hint">(військове звання, підпис, прізвище)</div>',
        '    </div>',
        '  </div>',
        '</div>',
        '</body></html>',
    ]

    log('Конвертую в PDF…')
    _html_to_pdf('\n'.join(L), out_path)
    log(f'✅  Збережено: {out_path.name}')


# ── Scale invoice ─────────────────────────────────────────────────────────────

def _nakladna_ws_to_html(ws) -> str:
    """Render openpyxl worksheet to HTML string, preserving merged cells."""
    from openpyxl.utils import get_column_letter

    merged_spans = {}
    covered = set()
    for rng in ws.merged_cells.ranges:
        r1, c1, r2, c2 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        merged_spans[(r1, c1)] = {'rowspan': r2 - r1 + 1, 'colspan': c2 - c1 + 1}
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if r != r1 or c != c1:
                    covered.add((r, c))

    col_widths = []
    for c in range(1, ws.max_column + 1):
        cd = ws.column_dimensions.get(get_column_letter(c))
        col_widths.append(cd.width if (cd and cd.width) else 8.0)
    total_w = sum(col_widths) or 1
    col_pcts = ''.join(
        f'<col style="width:{w / total_w * 100:.2f}%">' for w in col_widths
    )

    def _fmt(val):
        if val is None:
            return ''
        if isinstance(val, float):
            return f'{val:g}'
        return str(val)

    rows_html = []
    for r in range(1, ws.max_row + 1):
        cells = []
        for c in range(1, ws.max_column + 1):
            if (r, c) in covered:
                continue
            cell = ws.cell(r, c)
            val  = _fmt(cell.value)
            span_info = merged_spans.get((r, c), {})

            style_parts = []
            h_align = cell.alignment.horizontal if cell.alignment else None
            if h_align and h_align != 'general':
                style_parts.append(f'text-align:{h_align}')
            elif isinstance(cell.value, (int, float)) and cell.value is not None:
                style_parts.append('text-align:right')

            v_align = cell.alignment.vertical if cell.alignment else None
            if v_align:
                style_parts.append(f'vertical-align:{v_align}')

            wrap = cell.alignment.wrap_text if cell.alignment else False
            if not wrap:
                style_parts.append('white-space:nowrap')

            style_attr = f' style="{";".join(style_parts)}"' if style_parts else ''
            bold_open  = '<b>' if (cell.font and cell.font.bold) else ''
            bold_close = '</b>' if (cell.font and cell.font.bold) else ''

            attrs = style_attr
            if span_info.get('rowspan', 1) > 1:
                attrs += f' rowspan="{span_info["rowspan"]}"'
            if span_info.get('colspan', 1) > 1:
                attrs += f' colspan="{span_info["colspan"]}"'

            cells.append(f'<td{attrs}>{bold_open}{val}{bold_close}</td>')
        rows_html.append('<tr>' + ''.join(cells) + '</tr>')

    return (
        '<!DOCTYPE html><html lang="uk"><head><meta charset="utf-8"><style>'
        '@page{size:A4 portrait;margin:8mm 6mm}'
        'body{font-family:Arial,sans-serif;font-size:6.5pt}'
        'table{border-collapse:collapse;width:100%;table-layout:fixed}'
        'td{border:.4pt solid #000;padding:1pt 2pt;overflow:hidden;'
        'word-wrap:break-word;overflow-wrap:break-word}'
        '</style></head><body>'
        f'<table><colgroup>{col_pcts}</colgroup><tbody>'
        + ''.join(rows_html)
        + '</tbody></table></body></html>'
    )


def _amount_to_words(total_sum):
    grn = int(total_sum)
    kop = round((total_sum - grn) * 100)
    if kop == 100:
        grn += 1
        kop = 0
    words = num2words(grn, lang='uk', to='cardinal')
    return words, kop


def scale_nakladna(source_path: str, out_dir: str, base_count: int,
                   target_counts: list, progress_cb=None):
    """
    Scale invoice xlsx to each count in target_counts.
    base_count: original person count in source file (e.g. 520).
    target_counts: list of ints (e.g. [70, 450]).
    """
    def log(msg):
        if progress_cb:
            progress_cb(msg)

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    source_path = Path(source_path)
    stem = source_path.stem

    for target_count in target_counts:
        wb = openpyxl.load_workbook(str(source_path))
        ws = wb.active

        factor = target_count / base_count
        total_sum = 0.0

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            b_val = row[1].value
            if (b_val is not None and isinstance(b_val, (int, float))
                    and b_val == int(b_val) and int(b_val) > 0):

                price_cell = row[9]   # J
                sent_cell  = row[10]  # K
                recv_cell  = row[13]  # N
                sum_cell   = row[15]  # P

                if price_cell.value is None or sent_cell.value is None:
                    continue

                price   = float(price_cell.value)
                orig_q  = float(sent_cell.value)
                new_q   = round(orig_q * factor, 3)
                new_sum = round(new_q * price, 2)

                sent_cell.value = new_q
                if recv_cell.value is not None:
                    recv_cell.value = new_q
                if sum_cell.value is not None:
                    sum_cell.value = new_sum

                total_sum += new_sum

        total_sum = round(total_sum, 2)
        sum_words, kop = _amount_to_words(total_sum)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            d_val = row[3].value
            if d_val and isinstance(d_val, str) and 'тисяч' in d_val.lower():
                row[3].value = sum_words
                for cell in row:
                    if (cell.value and isinstance(cell.value, str)
                            and 'грн' in cell.value and 'коп' in cell.value):
                        cell.value = f'грн.   {kop:02d}   коп.'
                break

        new_stem = re.sub(r'\(\d+ ос\.\)', f'({target_count} ос.)', stem)
        if new_stem == stem:
            new_stem = f'{stem} ({target_count} ос.)'

        html = _nakladna_ws_to_html(ws)

        log(f'   [{target_count} ос.] Сума: {total_sum:.2f} грн  ({sum_words} грн. {kop:02d} коп.)')
        log('   Конвертую в PDF…')
        pdf_file = out_dir / f'{new_stem}.pdf'
        _html_to_pdf(html, pdf_file)
        log(f'✅  PDF: {pdf_file.name}')
