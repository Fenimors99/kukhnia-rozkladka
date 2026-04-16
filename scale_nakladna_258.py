import openpyxl
from num2words import num2words
import os
import copy

SOURCE = "/Users/vladandrieiev/Downloads/Telegram Desktop/Накладна 258 від 11.04.2026.xlsx"
OUTPUT_DIR = "/Users/vladandrieiev/Downloads/Telegram Desktop"
TOTAL_COUNT = 520


def amount_to_words(total_sum):
    """Convert sum to Ukrainian words + kopecks string."""
    grn = int(total_sum)
    kop = round((total_sum - grn) * 100)
    if kop == 100:
        grn += 1
        kop = 0
    words = num2words(grn, lang='uk', to='cardinal')
    return words, kop


def scale_nakladna(target_count):
    wb = openpyxl.load_workbook(SOURCE)
    ws = wb.active

    factor = target_count / TOTAL_COUNT
    total_sum = 0.0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        b_val = row[1].value  # col B — номер позиції
        if (b_val is not None
                and isinstance(b_val, (int, float))
                and b_val == int(b_val)
                and int(b_val) > 0):

            price_cell   = row[9]   # J — ціна
            sent_cell    = row[10]  # K — відправлено
            recv_cell    = row[13]  # N — прийнято
            sum_cell     = row[15]  # P — сума

            if price_cell.value is None or sent_cell.value is None:
                continue

            price   = float(price_cell.value)
            orig_q  = float(sent_cell.value)
            new_q   = round(orig_q * factor, 3)
            new_sum = round(new_q * price, 2)

            sent_cell.value = new_q
            if recv_cell.value is not None:
                recv_cell.value = new_q
            if sum_cell.value is not None:
                sum_cell.value = new_sum

            total_sum += new_sum

    total_sum = round(total_sum, 2)

    # Оновити суму прописом (рядок, де є "тисяч" або взагалі сума прописом у col D, index 3)
    sum_words, kop = amount_to_words(total_sum)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        d_val = row[3].value  # col D, index 3
        if d_val and isinstance(d_val, str) and 'тисяч' in d_val.lower():
            row[3].value = sum_words
            # Знайти "грн. XX коп." у тому ж рядку
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'грн' in cell.value and 'коп' in cell.value:
                    cell.value = f'грн.   {kop:02d}   коп.'
            break

    fname = f"Накладна 258 від 11.04.2026 ({target_count} ос.).xlsx"
    out_path = os.path.join(OUTPUT_DIR, fname)
    wb.save(out_path)
    print(f"[{target_count} ос.] Збережено: {out_path}")
    print(f"         Загальна сума: {total_sum:.2f} грн  ({sum_words} грн. {kop:02d} коп.)")


if __name__ == "__main__":
    for target in [70, 450]:
        scale_nakladna(target)
